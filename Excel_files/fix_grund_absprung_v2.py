"""
fix_grund_absprung_v2.py
------------------------
Zweiter Pass:
  1. Neue Eintraege zur GrundListe (Berechnung!J) hinzufuegen
  2. Benannten Bereich GrundAbsprung (und GrundListe) aktualisieren
  3. Erweiterte Mapping-Regeln anwenden (zweiter Pass auf Pipeline!O)
  4. Datenvalidierung sicherstellen
  5. Speichern & Report
"""

import re
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

FILE_PATH = "/Users/steffen/Documents/GitHub/Excel Leads/Excel_leads/Excel_files/Pipeline-Leads-26_04_18.xlsm"

# ---------------------------------------------------------------------------
# Neue Eintraege die zur GrundListe hinzugefuegt werden sollen
# ---------------------------------------------------------------------------
NEW_ENTRIES = [
    "Im Krankenhaus",
    "Ben\u00f6tigt 24h-Kraft",
    "Wollte sich nur informieren",
    "Noch kein Pflegegrad",
    "Weggezogen",
    "Doppelter Lead",
]

# ---------------------------------------------------------------------------
# Erweiterte explizite Mapping-Regeln (Reihenfolge: spezifischer zuerst)
# (norm_pattern -> ziel)
# ---------------------------------------------------------------------------
EXPLICIT_RULES = [
    # Doppelter Lead
    ("doppl",              "Doppelter Lead"),
    ("doppelt",            "Doppelter Lead"),
    # Im Krankenhaus
    ("im kh",              "Im Krankenhaus"),
    ("ins kh",             "Im Krankenhaus"),
    ("krankenhaus",        "Im Krankenhaus"),
    # Benoetigt 24h-Kraft
    ("24h",                "Ben\u00f6tigt 24h-Kraft"),
    ("24 stunden",         "Ben\u00f6tigt 24h-Kraft"),
    # Wollte sich nur informieren
    ("informier",          "Wollte sich nur informieren"),
    # Noch kein Pflegegrad
    ("noch kein pg",       "Noch kein Pflegegrad"),
    ("kein pg",            "Noch kein Pflegegrad"),
    # Hat sich erledigt
    ("hat sich erledigt",  "Hat sich erledigt"),
    ("schon versorgt",     "Hat sich erledigt"),
    ("erledigt",           "Hat sich erledigt"),
    # Anderer Dienst war schneller
    ("anderer dienst",     "Anderer Dienst war schneller"),
    ("alvital",            "Anderer Dienst war schneller"),
    # Pflegeheim
    ("pflegeheim",         "Pflegeheim"),
    # Noch kein Pflegegrad (pflegegrad als Fallback)
    ("pflegegrad",         "Noch kein Pflegegrad"),
    # Weggezogen
    ("weg gezogen",        "Weggezogen"),
    ("weggezogen",         "Weggezogen"),
    # Kein Bedarf
    ("psychose",           "Kein Bedarf"),
    # KH allein (nur als Wort, nicht Bestandteil eines anderen Worts)
    # Wird unten separat behandelt
]

# ---------------------------------------------------------------------------
# Hilfsfunktion: Normalisierung fuer Vergleiche
# ---------------------------------------------------------------------------
def normalize(s) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s

# ---------------------------------------------------------------------------
# Schritt 1 — Arbeitsmappe laden
# ---------------------------------------------------------------------------
print("Lade Arbeitsmappe ...")
wb = openpyxl.load_workbook(FILE_PATH, keep_vba=True, data_only=True)

ws_ber = wb["Berechnung"]

# Bestehende GrundListe lesen
grund_liste_existing = []
last_row = 1  # Header in Zeile 1
for row in ws_ber.iter_rows(min_row=2, max_row=200, min_col=10, max_col=10, values_only=True):
    val = row[0]
    last_row += 1
    if val is not None and str(val).strip():
        grund_liste_existing.append(str(val).strip())
    elif val is None:
        # Erste leere Zelle nach Daten = Ende der Liste
        # Aber wir merken uns last_row bereits oben
        break

# Genauer: wirklich letzte belegte Zeile finden
last_data_row = 1
for r in range(2, 200):
    cell = ws_ber.cell(row=r, column=10)
    if cell.value is not None and str(cell.value).strip():
        last_data_row = r

print(f"\nSchritt 1 — Bestehende GrundListe ({len(grund_liste_existing)} Eintraege, letzte Zeile: {last_data_row}):")
for g in grund_liste_existing:
    print(f"  '{g}'")

# ---------------------------------------------------------------------------
# Schritt 2 — Neue Eintraege hinzufuegen (nur falls nicht bereits vorhanden)
# ---------------------------------------------------------------------------
existing_norm = {normalize(g) for g in grund_liste_existing}
added_entries = []
next_row = last_data_row + 1

for entry in NEW_ENTRIES:
    if normalize(entry) not in existing_norm:
        ws_ber.cell(row=next_row, column=10).value = entry
        added_entries.append(entry)
        existing_norm.add(normalize(entry))
        next_row += 1
    else:
        print(f"  Bereits vorhanden, uebersprungen: '{entry}'")

# Vollstaendige, aktualisierte GrundListe
grund_liste = grund_liste_existing + added_entries
new_last_row = last_data_row + len(added_entries)

print(f"\nSchritt 2 — {len(added_entries)} neue Eintraege hinzugefuegt (Zeilen {last_data_row+1}..{new_last_row}):")
for e in added_entries:
    print(f"  + '{e}'")

# ---------------------------------------------------------------------------
# Schritt 3 — Benannte Bereiche aktualisieren
# ---------------------------------------------------------------------------
# GrundAbsprung und GrundListe sollen beide auf Berechnung!$J$2:$J$<new_last_row> zeigen
new_ref = f"Berechnung!$J$2:$J${new_last_row}"
print(f"\nSchritt 3 — Benannte Bereiche aktualisieren auf {new_ref} ...")

for name_to_update in ("GrundAbsprung", "GrundListe"):
    if name_to_update in wb.defined_names:
        dn = wb.defined_names[name_to_update]
        old_ref = dn.attr_text
        dn.attr_text = new_ref
        print(f"  {name_to_update}: '{old_ref}' -> '{new_ref}'")
    else:
        # Neu anlegen
        dn = DefinedName(name=name_to_update, attr_text=new_ref)
        wb.defined_names[name_to_update] = dn
        print(f"  {name_to_update}: NEU angelegt -> '{new_ref}'")

# ---------------------------------------------------------------------------
# Schritt 4 — Zweiter Mapping-Pass auf Pipeline!O
# ---------------------------------------------------------------------------
ws_pipe = wb["Pipeline"]

HEADER_ROW = 6
DATA_START  = 7
DATA_END    = 325

# Spaltenindex fuer "Grund zum Absprung" ermitteln
header_cells = list(ws_pipe.iter_rows(
    min_row=HEADER_ROW, max_row=HEADER_ROW,
    min_col=1, max_col=19, values_only=True
))[0]

col_idx = None
for i, h in enumerate(header_cells):
    if h is not None and "Grund zum Absprung" in str(h):
        col_idx = i + 1
        break

if col_idx is None:
    raise RuntimeError("Spalte 'Grund zum Absprung' nicht gefunden!")

col_letter = openpyxl.utils.get_column_letter(col_idx)
print(f"\nSchritt 4 — 'Grund zum Absprung' in Spalte {col_idx} ({col_letter})")

# GrundListe als Lookup
norm_map = {normalize(g): g for g in grund_liste}

def find_match_extended(raw: str):
    """
    Gibt (mapped_value, match_type) zurueck oder (None, None) wenn kein Match.
    match_type: 'exact', 'explicit', 'substring_gl_in_raw', 'substring_raw_in_gl'
    """
    norm_raw = normalize(raw)

    # "?" -> leer
    if norm_raw == "?":
        return ("", "explicit_empty")

    # 1. Exakter Match mit GrundListe
    if norm_raw in norm_map:
        return (norm_map[norm_raw], "exact")

    # 2. Explizite Regeln (in Reihenfolge)
    for pattern, target in EXPLICIT_RULES:
        if pattern in norm_raw:
            return (target, "explicit")

    # "kh" als eigenes Wort (nicht in "krankenhaus" etc.)
    if re.search(r'\bkh\b', norm_raw):
        return ("Im Krankenhaus", "explicit_kh")

    # 3. Substring: GrundListe-Eintrag enthalten in Rohwert (laengsten bevorzugen)
    candidates = []
    for norm_g, orig_g in norm_map.items():
        if norm_g and norm_g in norm_raw:
            candidates.append((len(norm_g), orig_g))
    if candidates:
        candidates.sort(key=lambda x: x[0], reverse=True)
        return (candidates[0][1], "substring_gl_in_raw")

    # 4. Substring: Rohwert enthalten in GrundListe-Eintrag
    for norm_g, orig_g in norm_map.items():
        if norm_raw and norm_raw in norm_g:
            return (orig_g, "substring_raw_in_gl")

    return (None, None)

# Alle Zellen lesen
print(f"\n  Lese O{DATA_START}:O{DATA_END} ...")
rohwerte = []
for r in range(DATA_START, DATA_END + 1):
    cell = ws_pipe.cell(row=r, column=col_idx)
    val = cell.value
    if val is not None and str(val).strip():
        rohwerte.append((r, str(val).strip()))

print(f"  {len(rohwerte)} nicht-leere Zellen gefunden.")

# Mapping durchfuehren
changed_count = 0
kein_match = []
already_valid = 0

print(f"\n  Mapping:")
for zeile, rohwert in rohwerte:
    norm_raw = normalize(rohwert)

    # Bereits exakter GrundListe-Eintrag -> unberuehrt lassen
    if norm_raw in norm_map and norm_map[norm_raw] == rohwert:
        already_valid += 1
        continue

    mapped, match_type = find_match_extended(rohwert)

    if mapped is None:
        kein_match.append((zeile, rohwert))
        print(f"    Zeile {zeile}: KEIN MATCH -> '{rohwert}'")
    elif mapped == rohwert:
        # Gleicher Wert, keine Aenderung noetig
        already_valid += 1
    else:
        cell = ws_pipe.cell(row=zeile, column=col_idx)
        cell.value = mapped
        changed_count += 1
        label = f"[{match_type}]"
        if mapped == "":
            print(f"    Zeile {zeile}: '{rohwert}' -> LEER {label}")
        else:
            print(f"    Zeile {zeile}: '{rohwert}' -> '{mapped}' {label}")

print(f"\n  Bereits gueltige Werte (unveraendert): {already_valid}")
print(f"  Geaenderte Zellen (zweiter Pass):      {changed_count}")
print(f"  Kein Match:                            {len(kein_match)}")

# ---------------------------------------------------------------------------
# Schritt 5 — Datenvalidierung sicherstellen
# ---------------------------------------------------------------------------
dv_range = f"{col_letter}{DATA_START}:{col_letter}{DATA_END}"
print(f"\nSchritt 5 — Datenvalidierung auf {dv_range} pruefen/setzen ...")

# Bestehende Validierungen auf dieser Spalte entfernen
to_remove = []
for dv in ws_pipe.data_validations.dataValidation:
    sqref_str = str(dv.sqref)
    if col_letter in sqref_str:
        to_remove.append(dv)
for dv in to_remove:
    ws_pipe.data_validations.dataValidation.remove(dv)
    print(f"  Bestehende Validierung entfernt: {dv.sqref}")

dv = DataValidation(
    type="list",
    formula1="GrundListe",
    showDropDown=False,
    showErrorMessage=True,
    errorStyle="warning",
    errorTitle="Unbekannter Wert",
    error="Bitte einen Wert aus der Liste w\u00e4hlen.",
    showInputMessage=False,
)
dv.sqref = dv_range
ws_pipe.add_data_validation(dv)
print(f"  Validierung gesetzt: formula1='GrundListe' auf {dv_range}")

# ---------------------------------------------------------------------------
# Schritt 6 — Speichern
# ---------------------------------------------------------------------------
print(f"\nSchritt 6 — Speichere Datei ...")
wb.save(FILE_PATH)
print(f"  Gespeichert: {FILE_PATH}")

# ---------------------------------------------------------------------------
# Zusammenfassung
# ---------------------------------------------------------------------------
print("\n" + "="*60)
print("ZUSAMMENFASSUNG")
print("="*60)
print(f"  GrundListe neue Eintraege hinzugefuegt:  {len(added_entries)}")
print(f"  GrundListe Gesamtgroesse:                {len(grund_liste)}")
print(f"  Zellen im zweiten Pass geaendert:        {changed_count}")
print(f"  Bereits gueltige Werte (unveraendert):   {already_valid}")
print(f"  Verbleibende KEIN-MATCH-Werte:           {len(kein_match)}")

if kein_match:
    print("\n  KEIN-MATCH-Werte (manuelle Pruefung erforderlich):")
    for zeile, val in kein_match:
        print(f"    Zeile {zeile}: '{val}'")
else:
    print("\n  Alle Werte konnten gemappt werden.")

print(f"\n  Datenvalidierung:  JA ({dv_range}, formula1='GrundListe')")
print("="*60)
