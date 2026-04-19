"""
fix_grund_absprung_v3.py
------------------------
Pipeline-Leads-26_04_19.xlsm:
  - Liest GrundAbsprung-Liste aus Berechnung!J
  - Mappt alle Werte in Pipeline[Grund zum Absprung] auf Listeneintraege
  - Explizite Regeln + Substring-Fallback
  - Stellt Datenvalidierung sicher
  - Speichert in-place
"""

import re
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

FILE_PATH = "/Users/steffen/Documents/GitHub/Excel Leads/Excel_leads/Excel_files/Pipeline-Leads-26_04_19.xlsm"

# ---------------------------------------------------------------------------
# Explizite Mapping-Regeln  (Reihenfolge: spezifischer zuerst)
# Muster (lowercase) -> Ziel-Eintrag aus GrundAbsprung
# ---------------------------------------------------------------------------
EXPLICIT_RULES = [
    # --- Doppelter Lead ---
    ("doppelt",             "Doppelter Lead"),
    ("doppl",               "Doppelter Lead"),
    ("dopplt",              "Doppelter Lead"),

    # --- Im Krankenhaus ---
    ("im kh",               "Im Krankenhaus"),
    ("ins kh",              "Im Krankenhaus"),
    ("krankenhaus",         "Im Krankenhaus"),

    # --- Benötigt 24h-Kraft ---
    ("24h kraft",           "Benötigt 24h-Kraft"),
    ("24h",                 "Benötigt 24h-Kraft"),
    ("24 stunden",          "Benötigt 24h-Kraft"),

    # --- Pflegeheim ---
    ("pflegeheim",          "Pflegeheim"),

    # --- Anderer Dienst war schneller ---
    ("alvital",             "Anderer Dienst war schneller"),
    ("anderer dienst",      "Anderer Dienst war schneller"),
    ("andere alternative",  "Sich für einen anderen Dienstleister entschieden"),
    ("alternative",         "Sich für einen anderen Dienstleister entschieden"),

    # --- Hat sich erledigt ---
    ("hat sich erledigt",   "Hat sich erledigt"),
    ("schon versorgt",      "Hat sich erledigt"),
    ("erledigt",            "Hat sich erledigt"),
    ("niemanden mehr",      "Hat sich erledigt"),
    ("benötigt niemanden",  "Hat sich erledigt"),

    # --- Noch kein Pflegegrad ---
    ("noch kein pg",        "Noch kein Pflegegrad"),
    ("kein pg",             "Noch kein Pflegegrad"),
    ("noch kein pflegegrad","Noch kein Pflegegrad"),
    ("kein pflegegrad",     "Noch kein Pflegegrad"),
    ("abwarten bis pg",     "Noch kein Pflegegrad"),
    ("kein interesse da noch", "Noch kein Pflegegrad"),

    # --- Kein Bedarf ---
    ("psychose",            "Kein Bedarf"),
    ("kein bedarf",         "Kein Bedarf"),
    ("aktuell kein bedarf", "Kein Bedarf"),

    # --- Falsches Gebiet ---
    ("rosenheim",           "Falsches Gebiet"),
    ("nürnberg",            "Falsches Gebiet"),
    ("ürnberg",             "Falsches Gebiet"),
    ("berlin",              "Falsches Gebiet"),
    ("braucht jemand",      "Falsches Gebiet"),
    ("falsches gebiet",     "Falsches Gebiet"),

    # --- Weggezogen ---
    ("weg gezogen",         "Weggezogen"),
    ("weggezogen",          "Weggezogen"),
    ("wohnt in",            "Weggezogen"),

    # --- Angehörige übernimmt selbst ---
    ("tochter übernimmt",   "Angehörige  übernimmt selbst"),
    ("sohn übernimmt",      "Angehörige  übernimmt selbst"),
    ("angehörige",          "Angehörige  übernimmt selbst"),

    # --- Hatte bereits jemanden ---
    ("hatte schon jemand",  "Hatte bereits jemanden"),
    ("hatte bereits",       "Hatte bereits jemanden"),

    # --- Möchte noch überlegen ---
    ("noch überlegen",      "Möchte noch überlegen"),
    ("möchte noch",         "Möchte noch überlegen"),
    ("noch nicht",          "Möchte noch überlegen"),
    ("erst wenn es schlechter", "Möchte noch überlegen"),

    # --- Kunde meldet sich bei Bedarf ---
    ("meldet sich eigenständig", "Kunde meldet sich bei Bedarf"),
    ("meldet sich von alleine",  "Kunde meldet sich bei Bedarf"),
    ("meldet sich wieder",       "meldet sich wieder"),
    ("meldet sich",              "Kunde meldet sich bei Bedarf"),
    ("meldet sich nochmal",      "Kunde meldet sich bei Bedarf"),

    # --- Kein weiterer Kontakt gewünscht ---
    ("kein weiterer anruf",    "Kein weiterer Kontakt gewünscht"),
    ("nicht mehr",             "Kein weiterer Kontakt gewünscht"),

    # --- Wollte sich nur informieren ---
    ("informier",           "Wollte sich nur informieren"),

    # --- Nicht erreicht ---
    ("nicht erreicht",      "Nicht erreicht"),

    # --- Nummer falsch ---
    ("nummer falsch",       "Nummer falsch"),
    ("nummer f.",           "Nummer falsch"),

    # --- Zu teuer ---
    ("zu teuer",            "Zu teuer"),
    ("war schon kunde",     "Zu teuer"),

    # --- Zu langsam ---
    ("zu langsam",          "Zu langsam"),

    # --- Sucht Tagespflege ---
    ("tagespflege",         "Sucht Tagespflege"),

    # --- Sucht betreutes Wohnen ---
    ("betreutes wohnen",    "Sucht betreutes Wohnen"),

    # --- Sich für anderen Dienstleister entschieden ---
    ("anderen dienstleister",   "Sich für einen anderen Dienstleister entschieden"),
    ("sucht pflegedienst",      "Sich für einen anderen Dienstleister entschieden"),
    ("benötigt pflegedienst",   "Sich für einen anderen Dienstleister entschieden"),
]


# ---------------------------------------------------------------------------
def normalize(s) -> str:
    if s is None:
        return ""
    s = str(s).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", s).strip().lower()


# ---------------------------------------------------------------------------
print("Lade Arbeitsmappe ...")
wb = openpyxl.load_workbook(FILE_PATH, keep_vba=True, data_only=True)

# --- GrundAbsprung-Liste aus Berechnung!J lesen ---
ws_ber = wb["Berechnung"]
grund_liste = []
last_data_row = 1
for r in range(2, 300):
    val = ws_ber.cell(row=r, column=10).value
    if val is None or not str(val).strip():
        break
    grund_liste.append(str(val).strip())
    last_data_row = r

print(f"\nGrundAbsprung-Liste ({len(grund_liste)} Eintraege, J2:J{last_data_row}):")
for g in grund_liste:
    print(f"  '{g}'")

# Benannten Bereich sicherstellen
ref_str = f"Berechnung!$J$2:$J${last_data_row}"
for name in ("GrundAbsprung", "GrundListe"):
    existing = wb.defined_names.get(name)
    if existing is not None:
        existing.attr_text = ref_str
        print(f"  Benannter Bereich '{name}' aktualisiert: {ref_str}")
    else:
        dn = DefinedName(name=name, attr_text=ref_str)
        wb.defined_names.append(dn)
        print(f"  Benannter Bereich '{name}' NEU: {ref_str}")

norm_map = {normalize(g): g for g in grund_liste}  # norm -> Original

# --- Spalte "Grund zum Absprung" in Pipeline lesen ---
ws_pipe = wb["Pipeline"]
HEADER_ROW = 6
DATA_START  = 7
DATA_END    = 500   # Sicherheitspuffer

header = list(ws_pipe.iter_rows(
    min_row=HEADER_ROW, max_row=HEADER_ROW, min_col=1, max_col=25, values_only=True
))[0]
col_idx = next(
    (i + 1 for i, h in enumerate(header) if h and "Grund zum Absprung" in str(h)), None
)
if col_idx is None:
    raise RuntimeError("Spalte 'Grund zum Absprung' nicht gefunden!")

col_letter = openpyxl.utils.get_column_letter(col_idx)
print(f"\nSpalte 'Grund zum Absprung': {col_letter} ({col_idx})")

# --- Tatsächliche letzte Datenzeile ermitteln ---
real_last = DATA_START - 1
for r in range(DATA_START, DATA_END + 1):
    # Irgendeine Zelle in dieser Zeile belegt?
    row_has_data = any(
        ws_pipe.cell(row=r, column=c).value is not None
        for c in range(1, 5)
    )
    if row_has_data:
        real_last = r

DATA_END = real_last
print(f"  Datenzeilen: {DATA_START} bis {DATA_END}")


# --- Matching-Funktion ---
def find_match(raw: str):
    """Gibt (mapped_value, match_type) zurück oder (None, None)."""
    norm_raw = normalize(raw)

    # Leer-Markierung
    if norm_raw in ("?", "-", ""):
        return ("", "leer")

    # 1. Exakter Match (case-insensitiv)
    if norm_raw in norm_map:
        return (norm_map[norm_raw], "exact")

    # 2. Explizite Regeln
    for pattern, target in EXPLICIT_RULES:
        if pattern in norm_raw:
            # Ziel muss in GrundListe existieren oder ist "" (leer)
            if target in grund_liste or target == "":
                return (target, "explicit")
            # Fallback: target als-is zurückgeben (Regel-Fehler abfangen)
            return (target, "explicit_notinlist")

    # "kh" als eigenes Wort
    if re.search(r'\bkh\b', norm_raw):
        return ("Im Krankenhaus", "explicit_kh")

    # 3. GrundListe-Eintrag als Substring im Rohwert
    candidates = [(len(ng), og) for ng, og in norm_map.items() if ng and ng in norm_raw]
    if candidates:
        candidates.sort(reverse=True)
        return (candidates[0][1], "substr_gl_in_raw")

    # 4. Rohwert als Substring in GrundListe-Eintrag
    for ng, og in norm_map.items():
        if norm_raw and norm_raw in ng:
            return (og, "substr_raw_in_gl")

    return (None, None)


# --- Mapping durchführen ---
rohwerte = []
for r in range(DATA_START, DATA_END + 1):
    val = ws_pipe.cell(row=r, column=col_idx).value
    if val is not None and str(val).strip():
        rohwerte.append((r, str(val).strip()))

print(f"\n{len(rohwerte)} nicht-leere Zellen gefunden.")
print("\nMapping:")

changed = 0
kein_match = []
already_ok = 0

for zeile, rohwert in rohwerte:
    norm_raw = normalize(rohwert)

    # Bereits korrekter Listenwert mit identischer Schreibweise
    if norm_raw in norm_map and norm_map[norm_raw] == rohwert:
        already_ok += 1
        continue

    mapped, mtype = find_match(rohwert)

    if mapped is None:
        kein_match.append((zeile, rohwert))
        print(f"  Zeile {zeile}: KEIN MATCH -> '{rohwert}'")
    elif mapped == rohwert:
        already_ok += 1
    else:
        ws_pipe.cell(row=zeile, column=col_idx).value = mapped
        changed += 1
        display = "LEER" if mapped == "" else f"'{mapped}'"
        print(f"  Zeile {zeile}: '{rohwert}' -> {display}  [{mtype}]")

print(f"\n  Bereits gueltig (unveraendert): {already_ok}")
print(f"  Geaendert:                      {changed}")
print(f"  Kein Match:                     {len(kein_match)}")

# --- Datenvalidierung ---
dv_range = f"{col_letter}{DATA_START}:{col_letter}{DATA_END}"
print(f"\nDatenvalidierung auf {dv_range} ...")

to_remove = [
    dv for dv in ws_pipe.data_validations.dataValidation
    if col_letter in str(dv.sqref)
]
for dv in to_remove:
    ws_pipe.data_validations.dataValidation.remove(dv)
    print(f"  Bestehende Validierung entfernt: {dv.sqref}")

new_dv = DataValidation(
    type="list",
    formula1="GrundListe",
    showDropDown=False,
    showErrorMessage=True,
    errorStyle="warning",
    errorTitle="Unbekannter Wert",
    error="Bitte einen Wert aus der Liste wählen.",
    showInputMessage=False,
)
new_dv.sqref = dv_range
ws_pipe.add_data_validation(new_dv)
print(f"  Validierung gesetzt: {dv_range} (formula1=GrundListe)")

# --- Speichern ---
print(f"\nSpeichere {FILE_PATH} ...")
wb.save(FILE_PATH)
print("Fertig.")

# --- Zusammenfassung ---
print("\n" + "=" * 60)
print("ZUSAMMENFASSUNG")
print("=" * 60)
print(f"  Nicht-leere Zellen gesamt:  {len(rohwerte)}")
print(f"  Bereits gueltig:            {already_ok}")
print(f"  Geaendert:                  {changed}")
print(f"  Kein Match (manuell):       {len(kein_match)}")

if kein_match:
    print("\n  KEIN-MATCH (manuelle Prüfung):")
    for z, v in kein_match:
        print(f"    Zeile {z}: '{v}'")
else:
    print("\n  Alle Werte gemappt.")
print("=" * 60)
