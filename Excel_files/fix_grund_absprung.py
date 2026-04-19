"""
fix_grund_absprung.py
---------------------
Liest GrundListe aus der Tabelle GrundAbsprung (Berechnung!J),
mappt alle Werte in Kundenliste[Grund zum Absprung] auf die
standardisierten Einträge und fügt eine Dropdown-Validierung hinzu.
"""

import re
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

FILE_PATH = "/Users/steffen/Documents/GitHub/Excel Leads/Excel_leads/Excel_files/Pipeline-Leads-26_04_18.xlsm"

# ---------------------------------------------------------------------------
# Hilfsfunktion: Normalisierung für Vergleiche
# ---------------------------------------------------------------------------
def normalize(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s

# ---------------------------------------------------------------------------
# Schritt 1 — GrundListe aus GrundAbsprung-Tabelle lesen
# ---------------------------------------------------------------------------
print("Lade Arbeitsmappe ...")
wb = openpyxl.load_workbook(FILE_PATH, keep_vba=True, data_only=True)

ws_ber = wb["Berechnung"]

# Tabelle GrundAbsprung liegt in J1:J25; Zeile 1 = Header "Grund zum Absprung"
grund_liste = []
for row in ws_ber.iter_rows(min_row=2, max_row=25, min_col=10, max_col=10, values_only=True):
    val = row[0]
    if val is not None and str(val).strip():
        grund_liste.append(str(val).strip())

print(f"\nSchritt 1 — GrundListe ({len(grund_liste)} Eintraege):")
for g in grund_liste:
    print(f"  '{g}'")

# ---------------------------------------------------------------------------
# Schritt 2 — Spalte "Grund zum Absprung" in Pipeline/Kundenliste lesen
# ---------------------------------------------------------------------------
ws_pipe = wb["Pipeline"]

# Tabelle Kundenliste: A6:S325 — Header in Zeile 6, Daten Zeilen 7..325
TABLE_REF   = "A6:S325"
HEADER_ROW  = 6
DATA_START  = 7
DATA_END    = 325

# Header-Zeile auslesen und Spaltenindex für "Grund zum Absprung" ermitteln
header_cells = list(ws_pipe.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW,
                                       min_col=1, max_col=19, values_only=True))[0]
col_idx = None  # 1-basiert
for i, h in enumerate(header_cells):
    if h is not None and "Grund zum Absprung" in str(h):
        col_idx = i + 1  # openpyxl-Spalte (1-basiert)
        break

if col_idx is None:
    raise RuntimeError("Spalte 'Grund zum Absprung' nicht gefunden!")

print(f"\nSchritt 2 — 'Grund zum Absprung' gefunden in Spalte {col_idx} "
      f"(Buchstabe: {openpyxl.utils.get_column_letter(col_idx)})")

# Alle nicht-leeren Datenzellen einlesen
rohwerte = []  # [(zeile_1basiert, rohwert)]
for r in range(DATA_START, DATA_END + 1):
    cell = ws_pipe.cell(row=r, column=col_idx)
    val = cell.value
    if val is not None and str(val).strip():
        rohwerte.append((r, str(val).strip()))

print(f"  {len(rohwerte)} nicht-leere Zellen gefunden.")

# ---------------------------------------------------------------------------
# Schritt 3 — Mapping
# ---------------------------------------------------------------------------
# Lookup-Dict: normalisierter GrundListe-Eintrag -> Original
norm_map = {normalize(g): g for g in grund_liste}

def find_best_match(raw: str) -> str | None:
    norm_raw = normalize(raw)

    # 1. Exact match
    if norm_raw in norm_map:
        return norm_map[norm_raw]

    # 2. Rohwert enthält GrundListe-Eintrag (Substring: GL-Eintrag in Rohwert)
    candidates = []
    for norm_g, orig_g in norm_map.items():
        if norm_g and norm_g in norm_raw:
            candidates.append((len(norm_g), orig_g))  # längstes bevorzugen
    if candidates:
        candidates.sort(key=lambda x: x[0], reverse=True)
        return candidates[0][1]

    # 3. GrundListe-Eintrag enthält Rohwert (Substring: Rohwert in GL-Eintrag)
    for norm_g, orig_g in norm_map.items():
        if norm_raw and norm_raw in norm_g:
            return orig_g

    return None

print(f"\nSchritt 3 — Mapping:")
mapping_results = []  # [(zeile, rohwert, mapped_value, changed)]
kein_match = []

for zeile, rohwert in rohwerte:
    best = find_best_match(rohwert)
    if best is None:
        mapping_results.append((zeile, rohwert, rohwert, False))
        kein_match.append((zeile, rohwert))
        print(f"  Zeile {zeile}: KEIN MATCH -> '{rohwert}'")
    elif best != rohwert:
        mapping_results.append((zeile, rohwert, best, True))
        print(f"  Zeile {zeile}: '{rohwert}' -> '{best}'")
    else:
        mapping_results.append((zeile, rohwert, best, False))

# ---------------------------------------------------------------------------
# Schritt 4 — Zurückschreiben
# ---------------------------------------------------------------------------
changed_count = 0
for zeile, rohwert, mapped, changed in mapping_results:
    if changed:
        cell = ws_pipe.cell(row=zeile, column=col_idx)
        cell.value = mapped
        changed_count += 1

print(f"\nSchritt 4 — {changed_count} Zellen wurden geaendert.")

# ---------------------------------------------------------------------------
# Schritt 5 — Datenvalidierung hinzufügen
# ---------------------------------------------------------------------------
col_letter = openpyxl.utils.get_column_letter(col_idx)
dv_range = f"{col_letter}{DATA_START}:{col_letter}{DATA_END}"

print(f"\nSchritt 5 — Fuege Dropdown-Validierung hinzu auf {dv_range} ...")

# Bestehende Validierungen auf dieser Spalte entfernen um Duplikate zu vermeiden
to_remove = []
for dv in ws_pipe.data_validations.dataValidation:
    # sqref ist ein MultiCellRange-Objekt; str() konvertieren für einfachen Check
    sqref_str = str(dv.sqref)
    if col_letter in sqref_str:
        to_remove.append(dv)
for dv in to_remove:
    ws_pipe.data_validations.dataValidation.remove(dv)
    print(f"  Bestehende Validierung entfernt: {dv.sqref}")

dv = DataValidation(
    type="list",
    formula1="GrundListe",
    showDropDown=False,       # False = Pfeil sichtbar
    showErrorMessage=True,
    errorStyle="warning",
    errorTitle="Unbekannter Wert",
    error="Bitte einen Wert aus der Liste w\u00e4hlen.",
    showInputMessage=False,
)
dv.sqref = dv_range
ws_pipe.add_data_validation(dv)
print(f"  Validierung erfolgreich hinzugefuegt: {dv_range}")
print(f"  formula1 = 'GrundListe'  |  showDropDown=False  |  errorStyle='warning'")

# ---------------------------------------------------------------------------
# Schritt 6 — Speichern
# ---------------------------------------------------------------------------
print(f"\nSchritt 6 — Speichere Datei ...")
wb.save(FILE_PATH)
print(f"  Gespeichert: {FILE_PATH}")

# ---------------------------------------------------------------------------
# Zusammenfassung
# ---------------------------------------------------------------------------
print("\n" + "="*60)
print("ZUSAMMENFASSUNG")
print("="*60)
print(f"  Gesamte nicht-leere Zellen in 'Grund zum Absprung': {len(rohwerte)}")
print(f"  Geaenderte Zellen (gemappt):                        {changed_count}")
print(f"  Kein Match (unveraendert):                          {len(kein_match)}")

if kein_match:
    print("\n  KEIN-MATCH-Werte (manuelle Pruefung erforderlich):")
    for zeile, val in kein_match:
        print(f"    Zeile {zeile}: '{val}'")
else:
    print("\n  Alle Werte konnten gemappt werden.")

print(f"\n  Datenvalidierung hinzugefuegt:                      JA ({dv_range})")
print("="*60)
